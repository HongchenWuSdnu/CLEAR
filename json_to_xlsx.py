#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CED_Dataset -> XLSX (build + clean + reorder top-20 comments + remove "转发微博" + drop rows)
Then call OpenAI API using Column B (news_text) and write the response to Column E.

Output columns:
  A: id
  B: news_text
  C: label (1=rumor, 0=non-rumor)
  D: empty
  E: llm_output (API result)
  F..: comments

Requirements:
  pip install openpyxl openai
  export OPENAI_API_KEY="YOUR_KEY"

Run:
  python ced_pipeline_minimal_api.py --dataset_dir CED_Dataset --output CED_final.xlsx
"""

import argparse
import os
import json
import re
import math
import unicodedata
import time
from datetime import datetime
from typing import Any, Dict, List, Optional, Set, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from openai import OpenAI


TOP_K_COMMENTS = 20
MAX_COMMENTS_PER_POST = 500
MMR_CANDIDATE_POOL = 200
MMR_NGRAM = 3
MMR_ALPHA = 0.75

API_MODEL_DEFAULT = "gpt-4o-mini"
API_TEMPERATURE = 0.2
API_MAX_OUTPUT_TOKENS = 512
API_SAVE_EVERY = 50
API_SLEEP_SECONDS = 0.0
API_MAX_RETRIES = 5

PROMPT = (
    "请分析所给微博短文的真实性。注意评估其内容的逻辑合理性，是否偏离科学常识或违背常规认知。"
    "评估措辞是否带有过度夸张、煽动性、情绪极化的语言风格。依据新闻主题和背景将其与类似的常规新闻比较，"
    "进而评估事件发生的时间、地点、主要人物和过程等细节信息是否模糊、其模糊是否合理。"
    "同时评估发布者的动机、是否存在立场偏倚，从而做出明智的判断。"
    "若判断为虚假新闻则返回“1”，判断为真实新闻则返回“0”，并返回该条新闻的可信度解释。"
)

SYSTEM_RULE = (
    "You are a careful factuality assessor. "
    "Output format:\n"
    "Line 1: a single digit 0 or 1 (0=true, 1=false).\n"
    "Line 2+: a brief credibility explanation.\n"
    "Do not add any extra prefixes."
)

FORWARD_RE = re.compile(r"^\s*转发微博\s*[。！？!?,，\.]*\s*$")


def extract_id(filename: str) -> str:
    base = os.path.basename(filename)
    if base.lower().endswith(".json"):
        base = base[:-5]
    if "_" in base:
        return base.split("_", 1)[0]
    m = re.match(r"(\d+)", base)
    return m.group(1) if m else base


def safe_load_json(path: str) -> Any:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def parse_dt(dt_str: str) -> Optional[datetime]:
    if not dt_str:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(dt_str, fmt)
        except Exception:
            pass
    try:
        s = dt_str.strip()
        if len(s) >= 19:
            return datetime.strptime(s[:19].replace("T", " "), "%Y-%m-%d %H:%M:%S")
    except Exception:
        return None
    return None


def is_forward_weibo(s: str) -> bool:
    if s is None:
        return False
    if not isinstance(s, str):
        s = str(s)
    return bool(FORWARD_RE.match(s))


def is_blank_cell(v) -> bool:
    if v is None:
        return True
    if isinstance(v, float):
        return math.isnan(v)
    return str(v).strip() == ""


URL_RE = re.compile(r"(https?://\S+|www\.\S+)", re.IGNORECASE)
MENTION_RE = re.compile(r"@[\w\-\u4e00-\u9fff]+")
BRACKET_EMOJI_RE = re.compile(r"\[[^\[\]]{1,20}\]")
ZWSP_RE = re.compile(r"[\u200b\u200c\u200d\ufeff]")
EMOJI_RE = re.compile(
    "["                       
    "\U0001F300-\U0001F5FF"
    "\U0001F600-\U0001F64F"
    "\U0001F680-\U0001F6FF"
    "\U0001F700-\U0001F77F"
    "\U0001F780-\U0001F7FF"
    "\U0001F800-\U0001F8FF"
    "\U0001F900-\U0001F9FF"
    "\U0001FA00-\U0001FAFF"
    "\U00002600-\U000026FF"
    "\U00002700-\U000027BF"
    "\U0000FE00-\U0000FE0F"
    "\U0000200D"
    "]+",
    flags=re.UNICODE
)


def is_pure_punct_or_space(s: str) -> bool:
    if s is None:
        return True
    s = str(s).strip()
    if not s:
        return True
    for ch in s:
        if ch.isspace():
            continue
        cat = unicodedata.category(ch)
        if not (cat.startswith("P") or cat.startswith("S")):
            return False
    return True


def clean_text(text: Optional[str]) -> str:
    if text is None:
        return ""
    if not isinstance(text, str):
        text = str(text)

    s = text.replace("\r\n", "\n").replace("\r", "\n")
    s = ZWSP_RE.sub("", s)

    s = s.replace("&amp;", "&").replace("&lt;", "<").replace("&gt;", ">") \
         .replace("&quot;", "\"").replace("&apos;", "'")

    s = URL_RE.sub(" ", s)
    s = MENTION_RE.sub(" ", s)
    s = BRACKET_EMOJI_RE.sub(" ", s)
    s = EMOJI_RE.sub(" ", s)

    s = re.sub(r"[ \t]+", " ", s)
    s = re.sub(r"\n{3,}", "\n\n", s)
    s = s.strip()

    if is_pure_punct_or_space(s):
        return ""

    s_no_punct = "".join(
        ch for ch in s if not unicodedata.category(ch).startswith(("P", "S"))
    ).strip()
    if not s_no_punct:
        return ""
    return s


WS_RE = re.compile(r"\s+")

def norm_for_sim(s: str) -> str:
    s = "" if s is None else str(s)
    s = s.strip()
    s = WS_RE.sub(" ", s)
    s = re.sub(r"[\u200b\u200c\u200d\ufeff]", "", s)
    return s

def grams(s: str, n: int) -> Set[str]:
    s = norm_for_sim(s).replace(" ", "")
    if not s:
        return set()
    if len(s) < n:
        return {s}
    return {s[i:i+n] for i in range(len(s) - n + 1)}

def jaccard(a: Set[str], b: Set[str]) -> float:
    if not a or not b:
        return 0.0
    inter = len(a & b)
    union = len(a | b)
    return inter / union if union else 0.0

def base_score(s: str) -> float:
    s = norm_for_sim(s).replace(" ", "")
    L = len(s)
    if L == 0:
        return 0.0
    uniq = len(set(s))
    diversity = min(1.0, uniq / max(1, L))
    return float(L) * (0.75 + 0.25 * diversity)

def pick_mmr_indices(comments: List[str], k: int) -> List[int]:
    n = len(comments)
    if n == 0:
        return []
    bases = [base_score(c) for c in comments]
    idx_all = [i for i in range(n) if bases[i] > 0]
    if not idx_all:
        return []
    idx_all.sort(key=lambda i: (-bases[i], i))
    cand = idx_all[:MMR_CANDIDATE_POOL] if len(idx_all) > MMR_CANDIDATE_POOL else idx_all

    gmap = {i: grams(comments[i], MMR_NGRAM) for i in cand}
    selected: List[int] = []
    remaining = set(cand)

    while len(selected) < k and remaining:
        best_i = None
        best_val = None
        for i in list(remaining):
            sim_penalty = 0.0
            if selected:
                gi = gmap.get(i, set())
                sim_penalty = max(jaccard(gi, gmap.get(j, set())) for j in selected)
            val = MMR_ALPHA * bases[i] - (1.0 - MMR_ALPHA) * sim_penalty * 100.0
            if best_val is None or val > best_val or (val == best_val and (best_i is None or i < best_i)):
                best_val = val
                best_i = i
        if best_i is None:
            break
        selected.append(best_i)
        remaining.remove(best_i)

    selected.sort()
    return selected


def collect_posts(original_dir: str) -> Dict[str, str]:
    posts: Dict[str, str] = {}
    if not os.path.isdir(original_dir):
        return posts
    for name in os.listdir(original_dir):
        if not name.lower().endswith(".json"):
            continue
        pid = extract_id(name)
        path = os.path.join(original_dir, name)
        try:
            obj = safe_load_json(path)
            if isinstance(obj, dict):
                posts[pid] = obj.get("text", "")
        except Exception:
            continue
    return posts


def collect_comments(folder: str) -> Dict[str, List[Tuple[Optional[datetime], str]]]:
    out: Dict[str, List[Tuple[Optional[datetime], str]]] = {}
    if not os.path.isdir(folder):
        return out
    for name in os.listdir(folder):
        if not name.lower().endswith(".json"):
            continue
        pid = extract_id(name)
        path = os.path.join(folder, name)
        try:
            arr = safe_load_json(path)
            if not isinstance(arr, list):
                continue
            lst = out.setdefault(pid, [])
            for item in arr:
                if not isinstance(item, dict):
                    continue
                text = item.get("text", "")
                dt = parse_dt(item.get("date", ""))
                lst.append((dt, text))
        except Exception:
            continue
    return out


def build_clean_filter_workbook(dataset_dir: str) -> Workbook:
    original_dir = os.path.join(dataset_dir, "original-microblog")
    rumor_dir = os.path.join(dataset_dir, "rumor-repost")
    nonrumor_dir = os.path.join(dataset_dir, "non-rumor-repost")

    posts = collect_posts(original_dir)
    rumor_comments = collect_comments(rumor_dir)
    nonrumor_comments = collect_comments(nonrumor_dir)

    labels: Dict[str, int] = {}
    comments_by_id: Dict[str, List[Tuple[Optional[datetime], str]]] = {}

    for pid, lst in rumor_comments.items():
        labels[pid] = 1
        comments_by_id.setdefault(pid, []).extend(lst)

    for pid, lst in nonrumor_comments.items():
        if pid not in labels:
            labels[pid] = 0
        comments_by_id.setdefault(pid, []).extend(lst)

    all_ids = set(posts.keys()) | set(comments_by_id.keys()) | set(labels.keys())

    def id_sort_key(x: str):
        return (0, int(x)) if x.isdigit() else (1, x)

    sorted_ids = sorted(all_ids, key=id_sort_key)

    comment_texts: Dict[str, List[str]] = {}
    max_n = 0
    for pid in sorted_ids:
        lst = comments_by_id.get(pid, [])
        lst_sorted = sorted(lst, key=lambda t: (t[0] is None, t[0] or datetime.max))
        texts = [t[1] for t in lst_sorted if isinstance(t[1], str) and t[1].strip() != ""]
        if MAX_COMMENTS_PER_POST is not None:
            texts = texts[:MAX_COMMENTS_PER_POST]
        comment_texts[pid] = texts
        max_n = max(max_n, len(texts))

    headers = ["id", "news_text", "label", "", ""] + [f"comment_{i+1}" for i in range(max_n)]

    wb = Workbook()
    ws = wb.active
    ws.title = "CED"

    ws.append(headers)
    header_font = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = wrap

    for pid in sorted_ids:
        row = [pid, posts.get(pid, ""), labels.get(pid, ""), "", ""]
        row.extend(comment_texts.get(pid, []))
        if len(row) < len(headers):
            row.extend([""] * (len(headers) - len(row)))
        ws.append(row)

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 80
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 55
    for i in range(6, min(len(headers), 6 + 20) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 45

    for r in range(2, ws.max_row + 1):
        ws.cell(r, 2).alignment = wrap
        ws.cell(r, 5).alignment = wrap
        for c in range(6, len(headers) + 1):
            ws.cell(r, c).alignment = wrap

    start_col = 6
    end_col = ws.max_column

    for r in range(2, ws.max_row + 1):
        ws.cell(r, 2).value = clean_text(ws.cell(r, 2).value)
        for c in range(start_col, end_col + 1):
            v = ws.cell(r, c).value
            if not is_blank_cell(v):
                ws.cell(r, c).value = clean_text(v)

    for r in range(2, ws.max_row + 1):
        comments: List[str] = []
        for c in range(start_col, end_col + 1):
            v = ws.cell(r, c).value
            if is_blank_cell(v):
                continue
            s = str(v).strip()
            if s:
                comments.append(s)
        if len(comments) > TOP_K_COMMENTS:
            sel_idx = pick_mmr_indices(comments, TOP_K_COMMENTS)
            if sel_idx:
                selected = [comments[i] for i in sel_idx]
                rest = [comments[i] for i in range(len(comments)) if i not in set(sel_idx)]
                new_comments = selected + rest
                for i, s in enumerate(new_comments):
                    col = start_col + i
                    if col > end_col:
                        break
                    ws.cell(r, col).value = s
                for c in range(start_col + len(new_comments), end_col + 1):
                    ws.cell(r, c).value = ""

    rows_to_delete: List[int] = []
    for r in range(2, ws.max_row + 1):
        comments: List[str] = []
        for c in range(start_col, end_col + 1):
            v = ws.cell(r, c).value
            if is_blank_cell(v):
                continue
            s = str(v).strip()
            if not s:
                continue
            if is_forward_weibo(s):
                continue
            comments.append(s)

        for i, s in enumerate(comments):
            col = start_col + i
            if col > end_col:
                break
            ws.cell(r, col).value = s
        for c in range(start_col + len(comments), end_col + 1):
            ws.cell(r, c).value = ""

        if len(comments) < TOP_K_COMMENTS:
            rows_to_delete.append(r)

    for r in reversed(rows_to_delete):
        ws.delete_rows(r, 1)

    return wb


def call_openai(client: OpenAI, model: str, news_text: str) -> str:
    user_text = f"{PROMPT}\n\n微博短文：\n{news_text}\n"
    resp = client.responses.create(
        model=model,
        input=[
            {"role": "system", "content": [{"type": "input_text", "text": SYSTEM_RULE}]},
            {"role": "user", "content": [{"type": "input_text", "text": user_text}]},
        ],
        temperature=API_TEMPERATURE,
        max_output_tokens=API_MAX_OUTPUT_TOKENS,
    )
    return (resp.output_text or "").strip()


def fill_col_e_with_api_and_save(wb: Workbook, model: str, output_path: str) -> None:
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        raise EnvironmentError("OPENAI_API_KEY is not set. Export OPENAI_API_KEY='...' first.")
    client = OpenAI(api_key=api_key)

    ws = wb["CED"]
    processed = 0
    written = 0
    errors = 0

    for r in range(2, ws.max_row + 1):
        news = ws.cell(r, 2).value
        if is_blank_cell(news):
            continue

        ok = False
        last_err = None
        for attempt in range(1, API_MAX_RETRIES + 1):
            try:
                out = call_openai(client, model, str(news).strip())
                ws.cell(r, 5).value = out
                ok = True
                written += 1
                break
            except Exception as e:
                last_err = e
                wait = min(60.0, (2 ** (attempt - 1)) * 1.5)
                print(f"[API] row={r} error attempt {attempt}/{API_MAX_RETRIES}: {e} -> backoff {wait:.1f}s")
                time.sleep(wait)

        if not ok:
            errors += 1
            ws.cell(r, 5).value = f"[ERROR] {type(last_err).__name__}: {last_err}"
            print(f"[API] row={r} failed; wrote error marker.")

        processed += 1
        if API_SLEEP_SECONDS > 0:
            time.sleep(API_SLEEP_SECONDS)

        if API_SAVE_EVERY > 0 and (processed % API_SAVE_EVERY == 0):
            wb.save(output_path)
            print(f"[SAVE] progress saved -> {output_path}")

    wb.save(output_path)
    print(f"[API] done: processed={processed}, written={written}, errors={errors}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dataset_dir", required=True, help="Path to CED_Dataset")
    ap.add_argument("--output", required=True, help="Output .xlsx")
    ap.add_argument("--model", default=API_MODEL_DEFAULT, help="OpenAI model name")
    args = ap.parse_args()

    wb = build_clean_filter_workbook(args.dataset_dir)
    wb.save(args.output)

    fill_col_e_with_api_and_save(wb, args.model, args.output)
    print(f"Saved final -> {args.output}")


if __name__ == "__main__":
    main()
